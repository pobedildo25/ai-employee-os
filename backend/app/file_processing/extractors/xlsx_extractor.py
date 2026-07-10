from openpyxl import load_workbook

from app.file_processing.interfaces.extractor import Extractor
from app.file_processing.models import ExtractedContent


class XlsxExtractor(Extractor):
    def extract(self, file_path: str) -> ExtractedContent:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheets: list[dict[str, object]] = []
        tables: list[dict[str, object]] = []
        text_parts: list[str] = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows: list[list[str]] = []
            for row in sheet.iter_rows(values_only=True):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                if any(value.strip() for value in row_values):
                    rows.append(row_values)
                    text_parts.append(" | ".join(row_values))

            sheets.append({"name": sheet_name, "row_count": len(rows)})
            tables.append({"sheet": sheet_name, "rows": rows})

        workbook.close()
        full_text = "\n".join(text_parts).strip()

        return ExtractedContent(
            text=full_text or None,
            metadata={"sheet_count": len(sheets), "format": "xlsx"},
            tables=tables,
            structure={"sheets": sheets},
        )
